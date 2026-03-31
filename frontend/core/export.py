"""
Exportação de dados para CSV compatível com SINAN.

O SINAN (Sistema de Informação de Agravos de Notificação) usa campos
padronizados. Este módulo mapeia os dados internos do NotificAI para
o layout de exportação mais próximo do SINAN NET.

Referência de campos: SINAN NET — Ficha de Notificação/Investigação
de Violência Doméstica, Sexual e/ou outras Violências (v2019).
"""

import csv
import io
import json
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import sqlite3

from core.database import get_cases_for_export


# ---------------------------------------------------------------------------
# Mapeamento NotificationType → código SINAN (campo 41 — tipo de violência)
# ---------------------------------------------------------------------------

_SINAN_TIPO_VIOLENCIA = {
    "Violência Física":             "1",
    "Violência Sexual":             "2",
    "Violência Psicológica/Moral":  "3",
    "Violência Autoprovocada":      "4",
    "Negligência/Abandono":         "5",
    "Trabalho Infantil":            "6",
    "Tráfico de Pessoas":           "7",
    "Outros/Não Classificado":      "9",
}

_SINAN_SEVERITY_TO_EVOLUCAO = {
    "CRÍTICO":       "3",  # Hospitalização
    "ALTO":          "2",  # Atendimento ambulatorial
    "MODERADO":      "2",
    "BAIXO":         "1",  # Alta / sem internação
    "MÍNIMO":        "1",
    "SEM INDICAÇÃO": "1",
}

# Cabeçalho SINAN simplificado (campos obrigatórios de notificação)
_SINAN_FIELDS = [
    "DT_NOTIF",        # Data de notificação
    "DT_OCOR",         # Data da ocorrência (document_date)
    "TP_NOT",          # Tipo de notificação (4 = violência)
    "ID_AGRAVO",       # CID-10 sugerido
    "TIPO_VIOL",       # Tipo de violência (código SINAN)
    "SEVERIDADE",      # Nível de severidade (NotificAI)
    "SCORE",           # Score de risco (NotificAI)
    "CONFIANCA",       # Confiança da classificação (0-1)
    "STATUS_CASO",     # Status do workflow
    "HASH_PACIENTE",   # Pseudônimo do paciente
    "QT_PAGINAS",      # Quantidade de páginas do documento
    "DT_ANALISE",      # Data/hora da análise
    "ID_ANALISE",      # ID interno da análise
]

# Sugestão de CID-10 por tipo (orientativo; deve ser confirmado pelo clínico)
_CID10_SUGERIDO = {
    "Violência Física":             "T74.1",
    "Violência Sexual":             "T74.2",
    "Violência Psicológica/Moral":  "T74.3",
    "Violência Autoprovocada":      "X84",
    "Negligência/Abandono":         "T74.0",
    "Trabalho Infantil":            "T74.4",
    "Tráfico de Pessoas":           "T74.8",
    "Outros/Não Classificado":      "T74.9",
}


def _format_date(iso_str: Optional[str]) -> str:
    """Converte ISO timestamp ou data parcial para DD/MM/YYYY."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        # Tenta só a parte da data
        try:
            return datetime.strptime(iso_str[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            return iso_str[:10]


def _row_to_sinan(row: Dict[str, Any]) -> Dict[str, str]:
    tipo = row.get("notification_type", "")
    return {
        "DT_NOTIF":      _format_date(row.get("analyzed_at")),
        "DT_OCOR":       _format_date(row.get("document_date")),
        "TP_NOT":        "4",
        "ID_AGRAVO":     _CID10_SUGERIDO.get(tipo, "T74.9"),
        "TIPO_VIOL":     _SINAN_TIPO_VIOLENCIA.get(tipo, "9"),
        "SEVERIDADE":    row.get("severity_level", ""),
        "SCORE":         f"{row.get('score', 0.0):.2f}",
        "CONFIANCA":     f"{row.get('confidence', 0.0):.2%}",
        "STATUS_CASO":   row.get("case_status", "pendente"),
        "HASH_PACIENTE": str(row.get("patient_hash") or "")[:16],
        "QT_PAGINAS":    str(row.get("page_count") or ""),
        "DT_ANALISE":    _format_date(row.get("analyzed_at")),
        "ID_ANALISE":    str(row.get("analysis_id", "")),
    }


def export_sinan_csv(
    conn: sqlite3.Connection,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    notification_type: Optional[str] = None,
) -> bytes:
    """
    Gera CSV no layout SINAN e retorna bytes prontos para download.

    Args:
        conn:              Conexão SQLite.
        start_date:        Data inicial (YYYY-MM-DD), opcional.
        end_date:          Data final (YYYY-MM-DD), opcional.
        notification_type: Filtro por tipo, opcional.

    Returns:
        Conteúdo CSV codificado em UTF-8-BOM (compatível com Excel).
    """
    rows = get_cases_for_export(conn, start_date, end_date, notification_type)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_SINAN_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(_row_to_sinan(row))
    return buf.getvalue().encode("utf-8-sig")  # BOM para Excel


def export_full_json(
    conn: sqlite3.Connection,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> bytes:
    """
    Exporta todos os casos em JSON estruturado (para pesquisadores / TI).
    Inclui probabilidades completas mas nenhum PII.
    """
    rows = get_cases_for_export(conn, start_date, end_date)
    export = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "total":       len(rows),
        "cases":       rows,
    }
    return json.dumps(export, ensure_ascii=False, indent=2).encode("utf-8")


def export_excel(
    conn: sqlite3.Connection,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Optional[bytes]:
    """
    Exporta planilha Excel com duas abas: resumo e detalhes.
    Requer openpyxl; retorna None se não instalado.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return None

    rows = get_cases_for_export(conn, start_date, end_date)
    wb = openpyxl.Workbook()

    # Aba 1 — Layout SINAN
    ws1 = wb.active
    ws1.title = "SINAN"
    ws1.append(_SINAN_FIELDS)
    for r in rows:
        mapped = _row_to_sinan(r)
        ws1.append([mapped.get(f, "") for f in _SINAN_FIELDS])

    # Aba 2 — Dados completos
    ws2 = wb.create_sheet("Completo")
    if rows:
        ws2.append(list(rows[0].keys()))
        for r in rows:
            ws2.append(list(r.values()))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
